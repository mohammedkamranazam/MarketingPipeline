"""
Acceptance Criteria:
- parse_tabular(data, mime_type) -> TabularParseResult extracts rows and headers from CSV/XLSX.
- CSV parser handles comma-separated UTF-8 files; returns headers (list[str]) and rows
  (list[dict[str, str]]).
- XLSX parser reads the first sheet; returns headers from row 1 and data from remaining rows.
- Unsupported mime_type raises ValueError with mime_type in the message.
- Empty files return TabularParseResult with empty headers and empty rows (not an error).
- Row values are always coerced to str (None cells become empty string "").
- TabularParseResult is a dataclass with headers (list[str]) and rows (list[dict[str, str]]).
"""

import csv
import io
from dataclasses import dataclass, field

SUPPORTED_TABULAR_MIME_TYPES = {
    "text/csv",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}


@dataclass
class TabularParseResult:
    headers: list[str] = field(default_factory=list)
    rows: list[dict[str, str]] = field(default_factory=list)


def parse_tabular(data: bytes, mime_type: str) -> TabularParseResult:
    if mime_type == "text/csv":
        return _parse_csv(data)
    if mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        return _parse_xlsx(data)
    raise ValueError(f"Unsupported tabular mime_type: {mime_type}")


def _parse_csv(data: bytes) -> TabularParseResult:
    text = data.decode("utf-8", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    headers = list(reader.fieldnames or [])
    rows = [{k: (v or "") for k, v in row.items()} for row in reader]
    return TabularParseResult(headers=headers, rows=rows)


def _parse_xlsx(data: bytes) -> TabularParseResult:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return TabularParseResult()

    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return TabularParseResult()

    headers = [str(cell) if cell is not None else "" for cell in all_rows[0]]
    rows: list[dict[str, str]] = []
    for raw_row in all_rows[1:]:
        row_dict = {
            headers[i]: (str(cell) if cell is not None else "")
            for i, cell in enumerate(raw_row)
            if i < len(headers)
        }
        rows.append(row_dict)
    return TabularParseResult(headers=headers, rows=rows)
